# app.py (Flask Backend)
from flask import Flask, render_template, request, jsonify
import csv
from pymongo import MongoClient
app = Flask(__name__)
client = MongoClient('mongodb://localhost:27017/') 

# Access or create a database
db = client['Credit']  
collection = db['Scores']  
import openpyxl
workbook = openpyxl.load_workbook("teacher.xlsx")
worksheet = workbook.active
filter = {"name": "abi"}

# Create a replacement document
document = {"name": "abi"}

# Upsert the document into the collection
result = collection.update_one(filter, {"$set": document}, upsert=True)
"""

# Print the updated document's ID
print("Updated document ID:", result.upserted_id)
"""
@app.route('/')
def index():
    aca=[1,2]
    aca[0]=int(worksheet[2][2].value)
    if(aca[0]<80):
         aca[0]=5
    if(aca[0]>80 and aca[0]<90):
         aca[0]=6
    if(aca[0]>=90 and aca[0]<95):
         aca[0]=7
    if(aca[0]>=95):
         aca[0]=10
    aca[1]=int(worksheet[2][3].value)
    if(aca[1]<80):
         aca[1]=5
    if(aca[1]>80 and aca[1]<90):
         aca[0]=6
    if(aca[1]>=90 and aca[1]<95):
         aca[1]=8
    if(aca[1]>=95):
         aca[1]=10
    acad=aca[0]+aca[1]

    res=[1,2,3,4]
    res[0]=int(worksheet[2][5].value)*3
    res[1]=int(worksheet[2][4].value)*2
    res[2]=int(worksheet[2][6].value)*4
    res[3]=int(worksheet[2][7].value)*6

    pub=res[0]+res[1]+res[2]+res[3]
    if(pub>15):
         pub=15

    ug=int(worksheet[2][8].value)*2
    pg=int(worksheet[2][9].value)*4
    phd=int(worksheet[2][10].value)*6

    gui=ug+pg+phd
    if(phd>15):
         gui=15

    ind=int(worksheet[2][11].value)*4
    if(ind>10):
         ind=10

    return render_template('card.html',acad=acad,pub=pub,gui=gui,ind=ind)

@app.route('/acasubmit', methods=['POST'])
def acasubmit():
    res = request.form.get('res')
    fee = request.form.get('fee')
    worksheet.cell(row=2, column=3, value=res)
    worksheet.cell(row=2, column=4, value=fee)
    workbook.save("teacher.xlsx")
    # Define the update operation to set the result field to 85
    update = {"$set": {"result": res}}
    collection.update_one(filter, update, upsert=True)
    update = {"$set": {"feedback": fee}}
    collection.update_one(filter, update, upsert=True)
    return render_template('result.html')

@app.route('/pubsubmit', methods=['POST'])
def pubsubmit():
    selected_option = request.form['publicationType']
    print(selected_option)
    publisherName = request.form.get('publisherName')
    title = request.form.get('title')
    date=request.form.get('date')
    pub = {"type": selected_option, "title":title ,"publisher":publisherName, "date": date}
    # Update the document
    existing_doc = collection.find_one(filter)
        # If "projects" field exists, push the new project
    if "projects" in existing_doc:
            collection.update_one(filter, {"$push": {"projects": pub}})
        # If "projects" field doesn't exist, create it with the new project
    else:
            collection.update_one(filter, {"$set": {"projects": [pub]}}, upsert=True)
    
    if(selected_option=="internationalConference"):
        val= worksheet[2][5].value
        val=val+1
        worksheet.cell(row=2, column=6, value=val)   
        workbook.save("teacher.xlsx")
    elif(selected_option=="nationalJournal"):
        val= worksheet[2][6].value
        val=val+1
        worksheet.cell(row=2, column=7, value=val)   
        workbook.save("teacher.xlsx")
    elif(selected_option=="internationalJournal"):
        val= worksheet[2][7].value
        val=val+1
        worksheet.cell(row=2, column=8, value=val)   
        workbook.save("teacher.xlsx")
    elif(selected_option=="nationalConference"):
        val= worksheet[2][4].value
        val=val+1
        worksheet.cell(row=2, column=5, value=val)   
        workbook.save("teacher.xlsx")

    return render_template("login.html")

@app.route('/guisubmit', methods=['POST'])
def guisubmit():
    selected_option = request.form['guidanceType']
    print(selected_option)
    title = request.form.get('projectTitle')
    date=request.form.get('year')
    pub = {"type": selected_option, "title":title ,"year": date}
    # Update the document
    existing_doc = collection.find_one(filter)
        # If "projects" field exists, push the new project
    if "guidance" in existing_doc:
            collection.update_one(filter, {"$addToSet": {"guidance": pub}})
        # If "projects" field doesn't exist, create it with the new project
    else:
            collection.update_one(filter, {"$set": {"guidance": [pub]}}, upsert=True)
    if(selected_option=="UG"):
        val= worksheet[2][8].value
        val=val+1
        worksheet.cell(row=2, column=9, value=val)   
        workbook.save("teacher.xlsx")
    elif(selected_option=="PG"):
        val= worksheet[2][9].value
        val=val+1
        worksheet.cell(row=2, column=10, value=val)   
        workbook.save("teacher.xlsx")
    elif(selected_option=="PhD"):
        val= worksheet[2][10].value
        val=val+1
        worksheet.cell(row=2, column=11, value=val)   
        workbook.save("teacher.xlsx")
    
    return render_template("login.html")

@app.route('/indsubmit', methods=['POST'])
def indsubmit():
    selected_option = request.form['collaborationType']
    print(selected_option)
    title = request.form.get('projectTitle')
    date=request.form.get('year')
    pub = {"type": selected_option, "title":title ,"date": date}
    # Update the document
    existing_doc = collection.find_one(filter)
        # If "projects" field exists, push the new project
    if "Industry" in existing_doc:
            collection.update_one(filter, {"$push": {"Industry": pub}})
        # If "projects" field doesn't exist, create it with the new project
    else:
            collection.update_one(filter, {"$set": {"Industry": [pub]}}, upsert=True)
    val=worksheet[2][11].value
    val=val+1
    worksheet.cell(row=2,column=12,value=val)

    return render_template("card.html")
    

@app.route('/card', methods=['POST'])
def dashboard():
    # Check login credentials (dummy authentication)
    username = request.form['username']
    password = request.form['password']
    
    if username == 'admin' and password == 'admin':
        return render_template('card.html',username == username)
    else:
        return render_template('login.html', message='Invalid credentials')

@app.route('/submit', methods=['POST'])
def submit():
    # Process form data and calculate total marks
    data = request.form.to_dict()

    # Calculate total marks
    academic_marks = int(data['results']) + int(data['feedbacks'])
    research_marks = (int(data['national_conference']) * 2) + (int(data['international_conference']) * 3) + \
                     (int(data['national_journal']) * 4) + (int(data['international_journal']) * 6)
    project_guidance_marks = (int(data['ug_projects']) * 2) + (int(data['pg_projects']) * 4) + \
                             (int(data['phd_projects']) * 6)
    industry_related_marks = (int(data['industry_projects']) * 4) + (int(data['consultancy']) * 4)
    faculty_organized_events_marks = int(data['organized_events']) * 3
    faculty_attended_events_marks = int(data['attended_events']) * 4

    administrative_duties = data['administrative_duties']
    administrative_marks = 0
    if administrative_duties in ['hod', 'club_incharge', 'lab_incharge', 'time_table_incharge', 'bus_incharge',
                                  'hostel_duty', 'iso_iqac']:
        administrative_marks = 5

    total_marks = academic_marks + research_marks + project_guidance_marks + industry_related_marks + \
                  faculty_organized_events_marks + faculty_attended_events_marks + administrative_marks

    # Write data to CSV
    with open('marks.csv', mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=data.keys())
        writer.writeheader()
        writer.writerow(data)

    return jsonify({'message': 'Marks submitted successfully', 'total_marks': total_marks})

if __name__ == '__main__':
    app.run(debug=True)
